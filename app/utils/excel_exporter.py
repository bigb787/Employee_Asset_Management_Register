"""Excel exports for assets, gatepass, and leavers."""
from __future__ import annotations

import json
from datetime import date
from io import BytesIO
from typing import Any, Optional

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import database
from app.utils.iso_audit import COUNTRY_KEYS, ISO_CONTROLS, country_from_location

HEADER_FILL = PatternFill("solid", fgColor="185FA5")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="F5F5F5")


def _today_str() -> str:
    return date.today().isoformat()


def _auto_width(ws, max_chars: int = 50) -> None:
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        letter = get_column_letter(col[0].column)
        maxlen = 0
        for cell in col:
            if cell.value is None:
                continue
            maxlen = max(maxlen, min(len(str(cell.value)), max_chars))
        ws.column_dimensions[letter].width = max(10, maxlen + 2)


def _style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = f"A{row + 1}"


def _zebra(ws, start_row: int = 2) -> None:
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row), start=0):
        if i % 2:
            for c in row:
                c.fill = ALT_FILL


def _iso_summary_rows() -> tuple[list[list[Any]], dict[str, int]]:
    rows_out: list[list[Any]] = []
    asset_order = [
        "laptops",
        "desktops",
        "monitors",
        "accessories",
        "networking",
        "cloud_assets",
        "infodesk_applications",
        "third_party_software",
        "ups",
        "mobile_phones",
        "scanners_printers",
        "cameras_dvr",
    ]
    country_totals = {k: 0 for k in COUNTRY_KEYS}
    for t in asset_order:
        recs = database.get_all(t)
        ctrl = ISO_CONTROLS.get(t, "")
        pii = sum(1 for r in recs if str(r.get("contains_pii") or "").strip() == "Yes")
        conf = sum(
            1
            for r in recs
            if str(r.get("iso_classification") or "").strip()
            in ("Confidential", "Restricted")
        )
        last_up = None
        for r in recs:
            u = r.get("updated_at") or r.get("created_at")
            if u and (last_up is None or str(u) > str(last_up)):
                last_up = u
        rows_out.append(
            [t, ctrl, len(recs), pii, conf, last_up or ""]
        )
        for r in recs:
            loc = (
                r.get("asset_location")
                if t
                in ("cloud_assets", "infodesk_applications", "third_party_software")
                else r.get("location")
            )
            cty = country_from_location(str(loc) if loc is not None else None)
            if cty:
                country_totals[cty] += 1
    return rows_out, country_totals


def _dump_sheet(ws, headers: list[str], data_rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    _style_header_row(ws, 1)
    _zebra(ws, 2)
    _auto_width(ws)


def export_all_assets(db: Any = None) -> StreamingResponse:
    _ = db
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws_iso = wb.create_sheet("ISO Summary", 0)
    hdr = [
        "Table Name",
        "ISO 27001 Control",
        "Total Records",
        "PII Records",
        "Confidential Records",
        "Last Updated",
    ]
    summary_rows, country_totals = _iso_summary_rows()
    ws_iso.append(hdr)
    for r in summary_rows:
        ws_iso.append(r)
    _style_header_row(ws_iso, 1)
    _zebra(ws_iso, 2)
    ws_iso.append([])
    ws_iso.append(["Country", "Total Assets"])
    for k in COUNTRY_KEYS:
        ws_iso.append([k, country_totals[k]])
    _auto_width(ws_iso)

    def dump_table(sheet_name: str, table: str, headers: list[str], keys: list[str]) -> None:
        ws = wb.create_sheet(sheet_name)
        recs = database.get_all(table)
        rows = [[r.get(k) for k in keys] for r in recs]
        _dump_sheet(ws, headers, rows)

    dump_table(
        "Laptops",
        "laptops",
        [
            "Asset Type",
            "Manufacturer",
            "Service Tag",
            "Model",
            "P/N",
            "Asset Owner",
            "Assigned To",
            "Asset Status",
            "Last Owner",
            "Dept",
            "Location",
            "Asset Health",
            "Warranty",
            "Install Date",
            "Date Added/Updated",
            "Processor",
            "RAM",
            "Hard Disk",
            "O/S",
            "Supt Vendor",
            "Keyboard",
            "Mouse",
            "Headphone",
            "USB Extender",
            "Contains PII",
            "ISO Class",
        ],
        [
            "asset_type",
            "asset_manufacturer",
            "service_tag",
            "model",
            "pn",
            "asset_owner",
            "assigned_to",
            "asset_status",
            "last_owner",
            "dept",
            "location",
            "asset_health",
            "warranty",
            "install_date",
            "date_added_updated",
            "processor",
            "ram",
            "hard_disk",
            "os",
            "supt_vendor",
            "keyboard",
            "mouse",
            "headphone",
            "usb_extender",
            "contains_pii",
            "iso_classification",
        ],
    )
    dump_table(
        "Desktops",
        "desktops",
        [
            "Asset Type",
            "Manufacturer",
            "Service Tag",
            "Model",
            "P/N",
            "Asset Owner",
            "Assigned To",
            "Asset Status",
            "Last Owner",
            "Dept",
            "Location",
            "Asset Health",
            "Warranty",
            "Install Date",
            "Date Added/Updated",
            "Processor",
            "O/S",
            "Supt Vendor",
            "Configuration",
            "Contains PII",
            "ISO Class",
        ],
        [
            "asset_type",
            "asset_manufacturer",
            "service_tag",
            "model",
            "pn",
            "asset_owner",
            "assigned_to",
            "asset_status",
            "last_owner",
            "dept",
            "location",
            "asset_health",
            "warranty",
            "install_date",
            "date_added_updated",
            "processor",
            "os",
            "supt_vendor",
            "configuration",
            "contains_pii",
            "iso_classification",
        ],
    )
    dump_table(
        "Monitors",
        "monitors",
        [
            "Asset Type",
            "Manufacturer",
            "Service Tag",
            "Model",
            "P/N",
            "Asset Owner",
            "Assigned To",
            "Asset Status",
            "Dept",
            "Location",
            "Asset Health",
            "Warranty",
            "Install Date",
            "Date Added/Updated",
            "Supt Vendor",
            "Contains PII",
            "ISO Class",
        ],
        [
            "asset_type",
            "asset_manufacturer",
            "service_tag",
            "model",
            "pn",
            "asset_owner",
            "assigned_to",
            "asset_status",
            "dept",
            "location",
            "asset_health",
            "warranty",
            "install_date",
            "date_added_updated",
            "supt_vendor",
            "contains_pii",
            "iso_classification",
        ],
    )
    dump_table(
        "Accessories",
        "accessories",
        [
            "Asset Type",
            "Manufacturer",
            "Model",
            "P/N",
            "Asset Owner",
            "Assigned To",
            "Asset Status",
            "Dept",
            "Location",
            "Warranty",
            "Install Date",
            "Date Added/Updated",
            "Supt Vendor",
            "Linked Device",
            "Contains PII",
            "ISO Class",
        ],
        [
            "asset_type",
            "asset_manufacturer",
            "model",
            "pn",
            "asset_owner",
            "assigned_to",
            "asset_status",
            "dept",
            "location",
            "warranty",
            "install_date",
            "date_added_updated",
            "supt_vendor",
            "linked_device_tag",
            "contains_pii",
            "iso_classification",
        ],
    )
    dump_table(
        "Networking",
        "networking",
        [
            "Asset Type",
            "Asset ID",
            "MAC ID",
            "Asset Owner",
            "Location",
            "Model",
            "S/N",
            "P/N",
            "Warranty",
            "Install Date",
            "O/S",
            "Supt Vendor",
            "Dept",
            "Configuration",
            "Contains PII",
            "Date Added/Updated",
            "ISO Class",
        ],
        [
            "asset_type",
            "asset_id",
            "mac_id",
            "asset_owner",
            "location",
            "model",
            "sn",
            "pn",
            "warranty",
            "install_date",
            "os",
            "supt_vendor",
            "dept",
            "configuration",
            "contains_pii",
            "date_added_updated",
            "iso_classification",
        ],
    )
    dump_table(
        "Cloud Assets",
        "cloud_assets",
        [
            "Asset",
            "Asset Type",
            "Asset Value",
            "Asset Owner",
            "Asset Location",
            "Contains PII",
            "Asset Region",
            "Date Added/Updated",
            "ISO Class",
        ],
        [
            "asset",
            "asset_type",
            "asset_value",
            "asset_owner",
            "asset_location",
            "contains_pii",
            "asset_region",
            "date_added_updated",
            "iso_classification",
        ],
    )
    dump_table(
        "Infodesk Applications",
        "infodesk_applications",
        [
            "Asset",
            "Asset Type",
            "Asset Value",
            "Asset Owner",
            "Asset Location",
            "Contains PII",
            "Date Added/Updated",
            "ISO Class",
        ],
        [
            "asset",
            "asset_type",
            "asset_value",
            "asset_owner",
            "asset_location",
            "contains_pii",
            "date_added_updated",
            "iso_classification",
        ],
    )
    dump_table(
        "Third Party Software",
        "third_party_software",
        [
            "Asset",
            "Asset Type",
            "Asset Value",
            "Asset Owner",
            "Asset Location",
            "Contains PII",
            "Date Added/Updated",
            "CVE Alert",
            "Setup",
            "Billing API",
            "Patch Status",
            "ISO Class",
        ],
        [
            "asset",
            "asset_type",
            "asset_value",
            "asset_owner",
            "asset_location",
            "contains_pii",
            "date_added_updated",
            "cve_alert",
            "setup",
            "billing_api",
            "patch_status",
            "iso_classification",
        ],
    )
    dump_table(
        "UPS",
        "ups",
        [
            "Asset Type",
            "Device ID",
            "Location",
            "Model",
            "Warranty",
            "Install Date",
            "Supt Vendor",
            "Dept",
            "Asset Owner",
            "Contains PII",
            "Date Added/Updated",
            "ISO Class",
        ],
        [
            "asset_type",
            "device_id",
            "location",
            "model",
            "warranty",
            "install_date",
            "supt_vendor",
            "dept",
            "asset_owner",
            "contains_pii",
            "date_added_updated",
            "iso_classification",
        ],
    )
    dump_table(
        "Mobile Phones",
        "mobile_phones",
        [
            "Asset Type",
            "Device ID",
            "Location",
            "Model",
            "P/N",
            "Warranty",
            "Supt Vendor",
            "Dept",
            "Asset Owner",
            "Contains PII",
            "Date Added/Updated",
            "ISO Class",
        ],
        [
            "asset_type",
            "device_id",
            "location",
            "model",
            "pn",
            "warranty",
            "supt_vendor",
            "dept",
            "asset_owner",
            "contains_pii",
            "date_added_updated",
            "iso_classification",
        ],
    )
    dump_table(
        "Scanners and Printers",
        "scanners_printers",
        [
            "Asset Type",
            "Device ID",
            "Location",
            "Model",
            "Service Tag",
            "P/N",
            "Warranty",
            "Supt Vendor",
            "Dept",
            "Description",
            "Asset Owner",
            "Contains PII",
            "Date Added/Updated",
            "ISO Class",
        ],
        [
            "asset_type",
            "device_id",
            "location",
            "model",
            "service_tag",
            "pn",
            "warranty",
            "supt_vendor",
            "dept",
            "description",
            "asset_owner",
            "contains_pii",
            "date_added_updated",
            "iso_classification",
        ],
    )
    dump_table(
        "Cameras and DVR",
        "cameras_dvr",
        [
            "Asset Type",
            "Location",
            "Invoice No",
            "Warranty",
            "Install Date",
            "Supt Vendor",
            "Dept",
            "Asset Owner",
            "Contains PII",
            "Date Added/Updated",
            "ISO Class",
        ],
        [
            "asset_type",
            "location",
            "invoice_no",
            "warranty",
            "install_date",
            "supt_vendor",
            "dept",
            "asset_owner",
            "contains_pii",
            "date_added_updated",
            "iso_classification",
        ],
    )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"ISO27001_Asset_Register_{_today_str()}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _format_gatepass_items(raw: Optional[str]) -> str:
    if not raw:
        return ""
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return raw
    if isinstance(data, list):
        parts = []
        for i, it in enumerate(data, start=1):
            if isinstance(it, dict):
                desc = it.get("description") or it.get("Description") or ""
                qty = it.get("qty") or it.get("Qty") or ""
                parts.append(f"{i}. {desc} (Qty: {qty})")
            else:
                parts.append(f"{i}. {it}")
        return " | ".join(parts)
    return str(data)


def export_gatepass(db: Any = None) -> StreamingResponse:
    _ = db
    wb = Workbook()
    ws = wb.active
    ws.title = "Gatepass"
    headers = [
        "Gatepass No",
        "Date",
        "Pass Type",
        "Issued To",
        "Person",
        "Dept Head",
        "Security Guard",
        "Expected Return Date",
        "Actual Return Date",
        "Status",
        "Remarks",
        "Items",
        "Created At",
        "Created By",
    ]
    ws.append(headers)
    fill_open = PatternFill("solid", fgColor="E6F1FB")
    font_open = Font(color="0C447C")
    fill_closed = PatternFill("solid", fgColor="EAF3DE")
    font_closed = Font(color="3B6D11")
    fill_cancel = PatternFill("solid", fgColor="FCEBEB")
    font_cancel = Font(color="A32D2D")

    gp_rows = database.get_all("gatepass")
    for r in gp_rows:
        ws.append(
            [
                r.get("gatepass_no"),
                r.get("gatepass_date"),
                r.get("pass_type"),
                r.get("issued_to"),
                r.get("person"),
                r.get("dept_head"),
                r.get("security_guard"),
                r.get("expected_return_date"),
                r.get("actual_return_date"),
                r.get("status"),
                r.get("remarks"),
                _format_gatepass_items(r.get("asset_items")),
                r.get("created_at"),
                r.get("created_by"),
            ]
        )

    _style_header_row(ws, 1)
    _zebra(ws, 2)
    col_status = headers.index("Status") + 1
    for row_idx, r in enumerate(gp_rows, start=2):
        st = str(r.get("status") or "").strip()
        cell = ws.cell(row=row_idx, column=col_status)
        if st == "Open":
            cell.fill = fill_open
            cell.font = font_open
        elif st == "Closed":
            cell.fill = fill_closed
            cell.font = font_closed
        elif st == "Cancelled":
            cell.fill = fill_cancel
            cell.font = font_cancel
    _auto_width(ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"Gatepass_Export_{_today_str()}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _sys_status_style(cell, val: str) -> None:
    v = str(val or "").strip()
    if v == "Revoked":
        cell.fill = PatternFill("solid", fgColor="EAF3DE")
        cell.font = Font(color="3B6D11")
    elif v == "Active":
        cell.fill = PatternFill("solid", fgColor="FCEBEB")
        cell.font = Font(color="A32D2D")
    else:
        cell.fill = PatternFill("solid", fgColor="F1EFE8")
        cell.font = Font(color="5F5E5A")


def export_leavers(db: Any = None) -> StreamingResponse:
    _ = db
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Leavers Summary"
    h1 = [
        "Employee Name",
        "Date of Leaving",
        "Dept",
        "Line Manager",
        "Email",
        "HW Handed Over",
        "Evidence Uploaded",
        "IT Peer Review",
        "Reporting Manager",
        "Confirmation Audit",
        "GitHub Ticket",
        "Overall Status",
        "Created At",
    ]
    ws1.append(h1)
    recs = database.get_all("leavers_checklist")
    sys_cols = [
        "email_groups",
        "infodesk_qa_dev",
        "infodesk_prod",
        "jira_and_wiki",
        "ms_office",
        "mongo_access",
        "azure_infodesk",
        "azure_wn_infodesk",
        "vpn",
        "wn_vpn",
        "azure_devops",
        "info_admin",
        "zabbix",
        "github",
        "infodesk_portal",
        "salesforce",
    ]

    for r in recs:
        ev = "Yes" if r.get("evidence_file_s3_key") else "No"
        ws1.append(
            [
                r.get("employee_name"),
                r.get("date_of_leaving"),
                r.get("department"),
                r.get("line_manager"),
                r.get("email_address"),
                r.get("hw_handed_over"),
                ev,
                r.get("it_peer_review"),
                r.get("reporting_manager"),
                r.get("confirmation_audit"),
                r.get("communication_github_ticket"),
                r.get("overall_status"),
                r.get("created_at"),
            ]
        )
    _style_header_row(ws1, 1)
    _zebra(ws1, 2)
    _auto_width(ws1)

    ws2 = wb.create_sheet("System Access Status")
    h2 = [
        "Employee Name",
        "Date of Leaving",
        "Email Groups",
        "Infodesk QA/Dev",
        "Infodesk Prod",
        "JIRA and Wiki",
        "MS Office",
        "Mongo Access",
        "Azure Infodesk",
        "Azure WN InfoDesk",
        "VPN",
        "WN VPN",
        "Azure Devops",
        "InfoAdmin",
        "Zabbix",
        "GitHub",
        "InfoDesk Portal",
        "Salesforce",
        "Completion %",
        "Overall Status",
    ]
    ws2.append(h2)

    def completion_pct(r: dict) -> float:
        done = sum(
            1
            for k in sys_cols
            if str(r.get(k) or "").strip() in ("Revoked", "N/A")
        )
        hw = 1 if str(r.get("hw_handed_over") or "").strip() == "Yes" else 0
        ev = 1 if r.get("evidence_file_s3_key") else 0
        total = len(sys_cols) + 2
        return round((done + hw + ev) / total * 100, 1) if total else 0.0

    for r in recs:
        row_idx = ws2.max_row + 1
        row_vals = [
            r.get("employee_name"),
            r.get("date_of_leaving"),
        ]
        for k in sys_cols:
            row_vals.append(r.get(k))
        pct = completion_pct(r)
        row_vals.extend([pct, r.get("overall_status")])
        ws2.append(row_vals)
        for ci, k in enumerate(sys_cols, start=3):
            _sys_status_style(ws2.cell(row=row_idx, column=ci), str(r.get(k) or ""))

    _style_header_row(ws2, 1)
    _zebra(ws2, 2)
    _auto_width(ws2)

    total = len(recs)
    revoked_all = sum(
        1
        for r in recs
        if all(str(r.get(k) or "").strip() == "Revoked" for k in sys_cols)
    )
    in_prog = sum(1 for r in recs if str(r.get("overall_status") or "") == "In Progress")
    on_hold = sum(1 for r in recs if str(r.get("overall_status") or "") == "On Hold")
    ws2.append([])
    ws2.append(
        [
            f"Total Leavers: {total}",
            f"Fully Revoked: {revoked_all}",
            f"In Progress: {in_prog}",
            f"On Hold: {on_hold}",
        ]
    )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"IT_Exit_Checklist_Export_{_today_str()}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
